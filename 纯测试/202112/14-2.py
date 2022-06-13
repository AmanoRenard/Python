from openpyxl import load_workbook
wb=load_workbook("hello.xlsx")
ws=wb.worksheets[0]
ws['A1'] = 10
ws.cell(row=2, column=1, value=20)
a=ws.max_row
b=ws.max_column
print(a,b)

for i in range(a):
    for o in range(b):
        print(ws.cell(row=i+1, column=o+1).value, end=' ')
    print()